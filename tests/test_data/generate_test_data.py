"""
Test Data Generator
Generates test data in multiple formats (Excel, JSON, Database).
"""

import os
import json
from pathlib import Path
from typing import Dict, List, Any

try:
    import openpyxl
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from sqlalchemy import create_engine, MetaData, Table, Column, Integer, String, Date
    HAS_SQLALCHEMY = True
except ImportError:
    HAS_SQLALCHEMY = False


class TestDataGenerator:
    """Generate test data in multiple formats."""

    # Sample test data
    MEMBERS_DATA = [
        {
            "id": 1,
            "first_name": "John",
            "last_name": "Doe",
            "designation": "Senior Trainer",
            "test_type": "happy_path",
            "description": "Valid member creation",
        },
        {
            "id": 2,
            "first_name": "Alice",
            "last_name": "Smith",
            "designation": "Junior Trainer",
            "test_type": "update_first",
            "description": "For updating first name",
        },
        {
            "id": 3,
            "first_name": "Bob",
            "last_name": "Johnson",
            "designation": "Trainer",
            "test_type": "update_last",
            "description": "For updating last name",
        },
        {
            "id": 4,
            "first_name": "Charlie",
            "last_name": "Brown",
            "designation": "Coordinator",
            "test_type": "update_designation",
            "description": "For updating designation",
        },
        {
            "id": 5,
            "first_name": "David",
            "last_name": "Wilson",
            "designation": "Lead Trainer",
            "test_type": "update_all",
            "description": "For updating all fields",
        },
        {
            "id": 6,
            "first_name": "Eve",
            "last_name": "Davis",
            "designation": "Assistant",
            "test_type": "delete_test",
            "description": "Member to be deleted",
        },
        {
            "id": 7,
            "first_name": "Frank",
            "last_name": "Miller",
            "designation": "Trainer",
            "test_type": "multiple_delete",
            "description": "First member for multi-delete test",
        },
        {
            "id": 8,
            "first_name": "Grace",
            "last_name": "Taylor",
            "designation": "Trainer",
            "test_type": "multiple_delete",
            "description": "Second member for multi-delete test",
        },
        {
            "id": 9,
            "first_name": "Henry",
            "last_name": "Anderson",
            "designation": "Trainer",
            "test_type": "multiple_delete",
            "description": "Third member for multi-delete test",
        },
        {
            "id": 10,
            "first_name": "Ivy",
            "last_name": "Thomas",
            "designation": "Manager",
            "test_type": "prepopulation",
            "description": "For testing form prepopulation",
        },
    ]

    COURSES_DATA = [
        {
            "id": 1,
            "course_name": "Python 101",
            "faculty_id": 1,
            "category": "Programming",
            "start_date": "2024-01-15",
            "end_date": "2024-03-15",
            "test_type": "happy_path",
            "description": "Intro to Python",
        },
        {
            "id": 2,
            "course_name": "Web Dev Basics",
            "faculty_id": 2,
            "category": "Web Dev",
            "start_date": "2024-02-01",
            "end_date": "2024-04-01",
            "test_type": "happy_path",
            "description": "HTML/CSS/JS",
        },
        {
            "id": 3,
            "course_name": "Data Science 101",
            "faculty_id": 3,
            "category": "Data Analysis",
            "start_date": "2024-01-20",
            "end_date": "2024-03-20",
            "test_type": "update_test",
            "description": "DS Fundamentals",
        },
        {
            "id": 4,
            "course_name": "DevOps Essentials",
            "faculty_id": 4,
            "category": "DevOps",
            "start_date": "2024-02-10",
            "end_date": "2024-04-10",
            "test_type": "update_test",
            "description": "CI/CD Pipeline",
        },
        {
            "id": 5,
            "course_name": "QA Testing",
            "faculty_id": 5,
            "category": "Testing",
            "start_date": "2024-03-01",
            "end_date": "2024-05-01",
            "test_type": "delete_test",
            "description": "Software Testing",
        },
    ]

    STUDENTS_DATA = [
        {
            "id": 1,
            "name": "Mark Johnson",
            "email": "mark.johnson@test.com",
            "course_id": 1,
            "skills": "Python, Django",
            "resume_url": "https://example.com/resumes/mark.pdf",
            "test_type": "happy_path",
            "description": "Student for Python course",
        },
        {
            "id": 2,
            "name": "Sarah Williams",
            "email": "sarah.williams@test.com",
            "course_id": 2,
            "skills": "HTML, CSS, JavaScript",
            "resume_url": "https://example.com/resumes/sarah.pdf",
            "test_type": "happy_path",
            "description": "Student for Web Dev course",
        },
        {
            "id": 3,
            "name": "Michael Brown",
            "email": "michael.brown@test.com",
            "course_id": 3,
            "skills": "Python, SQL, Tableau",
            "resume_url": "https://example.com/resumes/michael.pdf",
            "test_type": "update_test",
            "description": "Student for Data Science course",
        },
    ]

    @staticmethod
    def generate_excel(output_path: str = "tests/data/test_data.xlsx"):
        """Generate Excel test data file."""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        # Members sheet
        members_sheet = wb.create_sheet("Members")
        TestDataGenerator._write_sheet(members_sheet, TestDataGenerator.MEMBERS_DATA)

        # Courses sheet
        courses_sheet = wb.create_sheet("Courses")
        TestDataGenerator._write_sheet(courses_sheet, TestDataGenerator.COURSES_DATA)

        # Students sheet
        students_sheet = wb.create_sheet("Students")
        TestDataGenerator._write_sheet(students_sheet, TestDataGenerator.STUDENTS_DATA)

        wb.save(output_path)
        print(f"✓ Generated Excel test data: {output_path}")
        return output_path

    @staticmethod
    def generate_json(output_path: str = "tests/data/test_data.json"):
        """Generate JSON test data file."""
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        data = {
            "Members": TestDataGenerator.MEMBERS_DATA,
            "Courses": TestDataGenerator.COURSES_DATA,
            "Students": TestDataGenerator.STUDENTS_DATA,
        }

        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)

        print(f"✓ Generated JSON test data: {output_path}")
        return output_path

    @staticmethod
    def generate_database(db_url: str = "sqlite:///tests/data/test_data.db"):
        """Generate database with test data."""
        if not HAS_SQLALCHEMY:
            raise ImportError("sqlalchemy is required. Install with: pip install sqlalchemy")

        # Create directory if needed
        if db_url.startswith("sqlite:///"):
            db_path = db_url.replace("sqlite:///", "")
            os.makedirs(os.path.dirname(db_path) or ".", exist_ok=True)

        engine = create_engine(db_url)

        # This is simplified; in production you'd use proper ORM models
        with engine.connect() as conn:
            # Create tables (simplified SQL)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS members (
                    id INTEGER PRIMARY KEY,
                    first_name TEXT,
                    last_name TEXT,
                    designation TEXT,
                    test_type TEXT,
                    description TEXT
                )
            """)

            # Insert members data
            for member in TestDataGenerator.MEMBERS_DATA:
                conn.execute(f"""
                    INSERT INTO members VALUES (
                        {member['id']}, '{member['first_name']}', '{member['last_name']}',
                        '{member['designation']}', '{member['test_type']}', '{member['description']}'
                    )
                """)

            conn.commit()

        print(f"✓ Generated database test data: {db_url}")
        return db_url

    @staticmethod
    def _write_sheet(sheet, data: List[Dict[str, Any]]):
        """Write data to Excel sheet."""
        if not data:
            return

        # Write headers
        headers = list(data[0].keys())
        sheet.append(headers)

        # Write data rows
        for row in data:
            row_values = [row.get(header) for header in headers]
            sheet.append(row_values)


def generate_all_formats():
    """Generate test data in all available formats."""
    print("\n" + "=" * 70)
    print("GENERATING TEST DATA IN ALL FORMATS")
    print("=" * 70 + "\n")

    try:
        TestDataGenerator.generate_excel()
    except ImportError as e:
        print(f"⚠ Skipped Excel: {e}")

    TestDataGenerator.generate_json()

    try:
        TestDataGenerator.generate_database()
    except ImportError as e:
        print(f"⚠ Skipped Database: {e}")

    print("\n✓ Test data generation complete!")


if __name__ == "__main__":
    generate_all_formats()
