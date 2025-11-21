"""
Test Data Provider Architecture
Base class and implementations for different data sources (Excel, JSON, Database).
Supports pluggable data source handling for flexible test data management.
"""

from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional, Union
from enum import Enum
import os
import json
from pathlib import Path


class DataSourceType(Enum):
    """Enum for supported data sources."""
    EXCEL = "excel"
    JSON = "json"
    DATABASE = "database"


class TestDataProvider(ABC):
    """
    Abstract Base Class for Test Data Providers.
    All data providers must inherit from this class and implement required methods.
    """

    def __init__(self, source: str):
        """
        Initialize the test data provider.

        Args:
            source: Source path or connection string (file path or DB URL)
        """
        self.source = source
        self.data_cache = {}
        self.current_index = {}
        self._is_loaded = False

    @abstractmethod
    def load(self):
        """
        Load data from the source.
        Must be implemented by subclasses.
        """
        pass

    @abstractmethod
    def get_all_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Get all data from a specific sheet/table/collection.

        Args:
            sheet_name: Name of sheet/table/collection

        Returns:
            List of dictionaries containing all rows
        """
        pass

    @abstractmethod
    def get_next_data(self, sheet_name: str) -> Dict[str, Any]:
        """
        Get next row of data sequentially.

        Args:
            sheet_name: Name of sheet/table/collection

        Returns:
            Dictionary with next row's data
        """
        pass

    def get_by_id(self, sheet_name: str, id_value: Union[int, str]) -> Dict[str, Any]:
        """
        Get a specific row by ID.

        Args:
            sheet_name: Name of sheet/table/collection
            id_value: ID value to search for

        Returns:
            Dictionary with matching row data

        Raises:
            ValueError: If ID not found
        """
        data = self.get_all_data(sheet_name)
        for row in data:
            if row.get("id") == id_value or row.get("ID") == id_value:
                return row
        raise ValueError(f"No data found with ID: {id_value}")

    def get_by_name(self, sheet_name: str, name_value: str) -> Dict[str, Any]:
        """
        Get a specific row by name value.

        Args:
            sheet_name: Name of sheet/table/collection
            name_value: Name value to search for

        Returns:
            Dictionary with matching row data

        Raises:
            ValueError: If name not found
        """
        data = self.get_all_data(sheet_name)
        for row in data:
            for key in ["name", "first_name", "firstname", "Name", "FirstName"]:
                if row.get(key) == name_value:
                    return row
        raise ValueError(f"No data found with name: {name_value}")

    def get_by_field(self, sheet_name: str, field_name: str, field_value: Any) -> List[Dict[str, Any]]:
        """
        Get all rows matching a specific field value.

        Args:
            sheet_name: Name of sheet/table/collection
            field_name: Field name to search in
            field_value: Value to match

        Returns:
            List of matching rows

        Raises:
            ValueError: If field not found in any rows
        """
        data = self.get_all_data(sheet_name)
        results = [row for row in data if row.get(field_name) == field_value]

        if not results:
            raise ValueError(f"No data found with {field_name}={field_value}")

        return results

    def get_by_filter(self, sheet_name: str, **filters) -> List[Dict[str, Any]]:
        """
        Get rows matching multiple filter criteria.

        Args:
            sheet_name: Name of sheet/table/collection
            **filters: Field name and value pairs to match

        Returns:
            List of matching rows
        """
        data = self.get_all_data(sheet_name)
        results = []

        for row in data:
            if all(row.get(key) == value for key, value in filters.items()):
                results.append(row)

        return results

    def get_random_data(self, sheet_name: str) -> Dict[str, Any]:
        """
        Get a random row from a sheet.

        Args:
            sheet_name: Name of sheet/table/collection

        Returns:
            Random row as dictionary
        """
        import random
        data = self.get_all_data(sheet_name)
        if not data:
            raise ValueError(f"No data available in {sheet_name}")
        return random.choice(data)

    def reset_index(self, sheet_name: str = None):
        """
        Reset sequential index for a sheet or all sheets.

        Args:
            sheet_name: Sheet to reset. If None, resets all.
        """
        if sheet_name is None:
            self.current_index = {}
        else:
            self.current_index[sheet_name] = 0

    def get_sheet_names(self) -> List[str]:
        """
        Get list of available sheet/table/collection names.

        Returns:
            List of sheet names
        """
        return list(self.data_cache.keys())

    def get_row_count(self, sheet_name: str) -> int:
        """
        Get number of rows in a sheet.

        Args:
            sheet_name: Name of sheet/table/collection

        Returns:
            Number of rows
        """
        return len(self.get_all_data(sheet_name))

    def get_members_data(self) -> List[Dict[str, Any]]:
        """Get all members test data."""
        try:
            return self.get_all_data("Members")
        except ValueError:
            return []

    def get_courses_data(self) -> List[Dict[str, Any]]:
        """Get all courses test data."""
        try:
            return self.get_all_data("Courses")
        except ValueError:
            return []

    def get_students_data(self) -> List[Dict[str, Any]]:
        """Get all students test data."""
        try:
            return self.get_all_data("Students")
        except ValueError:
            return []

    def get_next_member(self) -> Dict[str, Any]:
        """Get the next member test data."""
        return self.get_next_data("Members")

    def get_next_course(self) -> Dict[str, Any]:
        """Get the next course test data."""
        return self.get_next_data("Courses")

    def get_next_student(self) -> Dict[str, Any]:
        """Get the next student test data."""
        return self.get_next_data("Students")

    def print_summary(self):
        """Print summary of all available test data."""
        print("\n" + "=" * 70)
        print("TEST DATA SUMMARY")
        print("=" * 70)
        print(f"Data Source Type: {self.__class__.__name__}")
        print(f"Source: {self.source}")
        print(f"Status: {'Loaded' if self._is_loaded else 'Not Loaded'}\n")

        for sheet_name in self.get_sheet_names():
            data = self.get_all_data(sheet_name)
            print(f"📋 Sheet: {sheet_name}")
            print(f"   Rows: {len(data)}")

            if data:
                print(f"   Columns: {list(data[0].keys())}")
                print(f"\n   Sample row:")
                for key, value in list(data[0].items())[:5]:
                    print(f"      {key}: {value}")
                print()


class ExcelTestDataProvider(TestDataProvider):
    """
    Excel-based Test Data Provider.
    Reads test data from Excel (.xlsx) files.
    """

    def __init__(self, excel_path: str):
        """
        Initialize Excel Test Data Provider.

        Args:
            excel_path: Path to Excel file
        """
        super().__init__(excel_path)
        self.workbook = None
        self.load()

    def load(self):
        """Load data from Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")

        if not os.path.exists(self.source):
            raise FileNotFoundError(f"Excel file not found: {self.source}")

        try:
            self.workbook = openpyxl.load_workbook(self.source)
            print(f"✓ Loaded Excel test data from: {self.source}")
            self._load_all_sheets()
            self._is_loaded = True
        except Exception as e:
            raise Exception(f"Failed to load Excel file: {e}")

    def _load_all_sheets(self):
        """Load all sheets from workbook into cache."""
        for sheet_name in self.workbook.sheetnames:
            self.data_cache[sheet_name] = self._read_sheet(sheet_name)

    def _read_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Read a sheet and convert to list of dictionaries.

        Args:
            sheet_name: Name of sheet to read

        Returns:
            List of dictionaries
        """
        sheet = self.workbook[sheet_name]
        data = []

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(cell.value)

        # Read data rows (skip header)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            row_dict = {}
            for col_idx, header in enumerate(headers):
                row_dict[header] = row[col_idx] if col_idx < len(row) else None

            data.append(row_dict)

        return data

    def get_all_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Get all data from a sheet."""
        if sheet_name not in self.data_cache:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {self.get_sheet_names()}")
        return self.data_cache[sheet_name]

    def get_next_data(self, sheet_name: str) -> Dict[str, Any]:
        """Get next row sequentially."""
        data = self.get_all_data(sheet_name)

        if sheet_name not in self.current_index:
            self.current_index[sheet_name] = 0

        if self.current_index[sheet_name] >= len(data):
            self.current_index[sheet_name] = 0

        row = data[self.current_index[sheet_name]]
        self.current_index[sheet_name] += 1
        return row


class JSONTestDataProvider(TestDataProvider):
    """
    JSON-based Test Data Provider.
    Reads test data from JSON files.
    """

    def __init__(self, json_path: str):
        """
        Initialize JSON Test Data Provider.

        Args:
            json_path: Path to JSON file
        """
        super().__init__(json_path)
        self.load()

    def load(self):
        """Load data from JSON file."""
        if not os.path.exists(self.source):
            raise FileNotFoundError(f"JSON file not found: {self.source}")

        try:
            with open(self.source, 'r') as f:
                json_data = json.load(f)

            # Support both flat structure and nested by sheet names
            if isinstance(json_data, dict):
                self.data_cache = {
                    key: value if isinstance(value, list) else [value]
                    for key, value in json_data.items()
                }
            else:
                raise ValueError("JSON file must contain a dictionary with sheet names as keys")

            print(f"✓ Loaded JSON test data from: {self.source}")
            self._is_loaded = True
        except json.JSONDecodeError as e:
            raise Exception(f"Invalid JSON file: {e}")
        except Exception as e:
            raise Exception(f"Failed to load JSON file: {e}")

    def get_all_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Get all data from a sheet."""
        if sheet_name not in self.data_cache:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {self.get_sheet_names()}")
        return self.data_cache[sheet_name]

    def get_next_data(self, sheet_name: str) -> Dict[str, Any]:
        """Get next row sequentially."""
        data = self.get_all_data(sheet_name)

        if sheet_name not in self.current_index:
            self.current_index[sheet_name] = 0

        if self.current_index[sheet_name] >= len(data):
            self.current_index[sheet_name] = 0

        row = data[self.current_index[sheet_name]]
        self.current_index[sheet_name] += 1
        return row


class DBTestDataProvider(TestDataProvider):
    """
    Database-based Test Data Provider.
    Reads test data from a SQL database (SQLite, PostgreSQL, MySQL, etc.).
    """

    def __init__(self, connection_string: str, tables: Optional[Dict[str, str]] = None):
        """
        Initialize Database Test Data Provider.

        Args:
            connection_string: SQLAlchemy connection string
                              Examples:
                              - sqlite:///test_data.db
                              - postgresql://user:password@localhost/db
                              - mysql+pymysql://user:password@localhost/db
            tables: Dictionary mapping sheet names to table names
                   Example: {"Members": "members_test_data", "Courses": "courses_test_data"}
        """
        super().__init__(connection_string)
        self.tables = tables or {}
        self.engine = None
        self.connection = None
        self.load()

    def load(self):
        """Load data from database."""
        try:
            from sqlalchemy import create_engine, inspect
            from sqlalchemy import text
        except ImportError:
            raise ImportError(
                "sqlalchemy is required for database support. Install with: pip install sqlalchemy"
            )

        try:
            self.engine = create_engine(self.source)
            self.connection = self.engine.connect()

            # Get table names from database if not provided
            inspector = inspect(self.engine)
            available_tables = inspector.get_table_names()

            if not self.tables:
                # Use all available tables
                self.tables = {table: table for table in available_tables}

            # Load data from each table
            for sheet_name, table_name in self.tables.items():
                if table_name not in available_tables:
                    print(f"⚠ Warning: Table '{table_name}' not found in database")
                    continue

                self.data_cache[sheet_name] = self._read_table(table_name)

            print(f"✓ Loaded database test data from: {self.source}")
            self._is_loaded = True
        except Exception as e:
            raise Exception(f"Failed to load database: {e}")

    def _read_table(self, table_name: str) -> List[Dict[str, Any]]:
        """
        Read a table and convert to list of dictionaries.

        Args:
            table_name: Name of table to read

        Returns:
            List of dictionaries
        """
        try:
            from sqlalchemy import text
        except ImportError:
            raise ImportError("sqlalchemy is required")

        query = text(f"SELECT * FROM {table_name}")
        result = self.connection.execute(query)

        data = []
        for row in result:
            row_dict = dict(row)
            data.append(row_dict)

        return data

    def get_all_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Get all data from a table."""
        if sheet_name not in self.data_cache:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {self.get_sheet_names()}")
        return self.data_cache[sheet_name]

    def get_next_data(self, sheet_name: str) -> Dict[str, Any]:
        """Get next row sequentially."""
        data = self.get_all_data(sheet_name)

        if sheet_name not in self.current_index:
            self.current_index[sheet_name] = 0

        if self.current_index[sheet_name] >= len(data):
            self.current_index[sheet_name] = 0

        row = data[self.current_index[sheet_name]]
        self.current_index[sheet_name] += 1
        return row

    def close(self):
        """Close database connection."""
        if self.connection:
            self.connection.close()
        if self.engine:
            self.engine.dispose()

    def __del__(self):
        """Ensure connection is closed on object deletion."""
        self.close()
