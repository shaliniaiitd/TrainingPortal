"""
Data Validation Runner - validation_runner.py

Run validations on test data from multiple sources (JSON, Excel, Database).
Provides comprehensive validation reports.
"""

import json
import sys
from pathlib import Path
from typing import Dict, Any, List, Optional

from validators import (
    TestDataValidator,
    MembersValidator,
    CoursesValidator,
    StudentsValidator,
    ValidationResult
)


class DataLoader:
    """Load test data from different sources."""
    
    @staticmethod
    def load_json(filepath: str) -> Dict[str, List[Dict[str, Any]]]:
        """Load test data from JSON file."""
        with open(filepath, 'r') as f:
            data = json.load(f)
        return data
    
    @staticmethod
    def load_excel(filepath: str) -> Dict[str, List[Dict[str, Any]]]:
        """Load test data from Excel file."""
        try:
            import openpyxl
        except ImportError:
            print("❌ openpyxl not installed. Install with: pip install openpyxl")
            return {}
        
        data = {}
        wb = openpyxl.load_workbook(filepath)
        
        for sheet_name in ["Members", "Courses", "Students"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                    rows.append(row_dict)
                
                data[sheet_name] = rows
        
        return data


class ValidationRunner:
    """Run comprehensive data validations."""
    
    @classmethod
    def validate_json_file(cls, filepath: str) -> Dict[str, ValidationResult]:
        """Validate JSON test data file."""
        print(f"\n📄 Validating JSON file: {filepath}")
        
        try:
            data = DataLoader.load_json(filepath)
        except FileNotFoundError:
            print(f"❌ File not found: {filepath}")
            return {}
        except json.JSONDecodeError as e:
            print(f"❌ Invalid JSON: {e}")
            return {}
        
        members = data.get("Members", [])
        courses = data.get("Courses", [])
        students = data.get("Students", [])
        
        return TestDataValidator.validate_all(members, courses, students)
    
    @classmethod
    def validate_excel_file(cls, filepath: str) -> Dict[str, ValidationResult]:
        """Validate Excel test data file."""
        print(f"\n📊 Validating Excel file: {filepath}")
        
        try:
            data = DataLoader.load_excel(filepath)
        except FileNotFoundError:
            print(f"❌ File not found: {filepath}")
            return {}
        except Exception as e:
            print(f"❌ Error reading Excel: {e}")
            return {}
        
        if not data:
            print("❌ No data loaded from Excel file")
            return {}
        
        members = data.get("Members", [])
        courses = data.get("Courses", [])
        students = data.get("Students", [])
        
        return TestDataValidator.validate_all(members, courses, students)
    
    @classmethod
    def validate_all_sources(cls) -> None:
        """Validate all available test data sources."""
        print("\n" + "=" * 70)
        print("COMPREHENSIVE DATA VALIDATION")
        print("=" * 70)
        
        test_data_dir = Path("tests/test_data")
        
        all_results = {}
        
        # Check JSON
        json_file = test_data_dir / "test_data.json"
        if json_file.exists():
            results = cls.validate_json_file(str(json_file))
            all_results["JSON"] = results
            TestDataValidator.print_report(results)
        else:
            print(f"⚠️  JSON file not found: {json_file}")
        
        # Check Excel
        xlsx_file = test_data_dir / "test_data.xlsx"
        if xlsx_file.exists():
            results = cls.validate_excel_file(str(xlsx_file))
            all_results["Excel"] = results
            TestDataValidator.print_report(results)
        else:
            print(f"⚠️  Excel file not found: {xlsx_file}")
        
        # Summary
        cls._print_summary(all_results)
    
    @classmethod
    def _print_summary(cls, all_results: Dict[str, Dict[str, ValidationResult]]) -> None:
        """Print summary of all validations."""
        print("\n" + "=" * 70)
        print("VALIDATION SUMMARY")
        print("=" * 70 + "\n")
        
        for source_name, results in all_results.items():
            print(f"{source_name}:")
            for entity_name, result in results.items():
                status = "✅" if result.is_valid else "❌"
                error_count = len(result.errors)
                warning_count = len(result.warnings)
                print(f"  {status} {entity_name}: {error_count} errors, {warning_count} warnings")
        
        print("\n" + "=" * 70)


def main():
    """Main entry point."""
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        if filepath.endswith(".json"):
            results = ValidationRunner.validate_json_file(filepath)
        elif filepath.endswith(".xlsx"):
            results = ValidationRunner.validate_excel_file(filepath)
        else:
            print("❌ Unsupported file format. Use .json or .xlsx")
            return 1
        
        TestDataValidator.print_report(results)
        
        # Return 0 if all valid, 1 if any errors
        all_valid = all(r.is_valid for r in results.values())
        return 0 if all_valid else 1
    else:
        ValidationRunner.validate_all_sources()
        return 0


if __name__ == "__main__":
    sys.exit(main())
