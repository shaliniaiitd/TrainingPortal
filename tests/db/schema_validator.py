"""
Schema Validator - schema_validator.py

Validates test data structure and format consistency across all sources.
Ensures JSON, Excel, and database files follow the expected schema.
This version adds direct SQLite database validation (validate_db_structure).
"""

import json
import sqlite3
from typing import Dict, List, Any, Tuple, Optional
from pathlib import Path
from dataclasses import dataclass


@dataclass
class SchemaValidationResult:
    """Result of schema validation."""
    is_valid: bool
    errors: List[str]
    warnings: List[str]
    
    def __str__(self) -> str:
        """Format result as string."""
        status = "✅ VALID" if self.is_valid else "❌ INVALID"
        lines = [f"Status: {status}"]
        
        if self.errors:
            lines.append(f"Errors ({len(self.errors)}):")
            for error in self.errors:
                lines.append(f"  ❌ {error}")
        
        if self.warnings:
            lines.append(f"Warnings ({len(self.warnings)}):")
            for warning in self.warnings:
                lines.append(f"  ⚠️  {warning}")
        
        return "\n".join(lines)


class SchemaDefinition:
    """Schema definitions for each entity type."""
    
    # Members schema
    MEMBERS_SCHEMA = {
        "required_fields": ["id", "first_name", "last_name", "designation"],
        "optional_fields": ["email", "phone", "department"],
        "field_types": {
            "id": (int, "unique identifier"),
            "first_name": (str, "first name"),
            "last_name": (str, "last name"),
            "designation": (str, "job designation"),
            "email": (str, "email address"),
            "phone": (str, "phone number"),
            "department": (str, "department name")
        },
        "field_lengths": {
            "first_name": (2, 100),
            "last_name": (2, 100),
            "designation": (3, 50),
            "email": (5, 100),
            "phone": (7, 20),
            "department": (2, 50)
        },
        "valid_enums": {
            "designation": [
                "Professor", "Associate Professor", "Assistant Professor",
                "Lecturer", "Senior Lecturer", "Instructor", "Teaching Assistant"
            ]
        }
    }
    
    # Courses schema
    COURSES_SCHEMA = {
        "required_fields": ["id", "course_name", "faculty_id", "category"],
        "optional_fields": ["start_date", "end_date", "description", "duration_hours"],
        "field_types": {
            "id": (int, "unique identifier"),
            "course_name": (str, "course name"),
            "faculty_id": (int, "faculty member ID"),
            "category": (str, "course category"),
            "start_date": (str, "start date (YYYY-MM-DD)"),
            "end_date": (str, "end date (YYYY-MM-DD)"),
            "description": (str, "course description"),
            "duration_hours": (int, "duration in hours")
        },
        "field_lengths": {
            "course_name": (3, 100),
            "category": (3, 50),
            "description": (0, 500)
        },
        "valid_enums": {
            "category": [
                "Programming", "Web Development", "Data Analysis",
                "DevOps", "Testing", "Others"
            ]
        }
    }
    
    # Students schema
    STUDENTS_SCHEMA = {
        "required_fields": ["id", "name", "email", "course_id"],
        "optional_fields": ["resume", "skills", "enrollment_date"],
        "field_types": {
            "id": (int, "unique identifier"),
            "name": (str, "student name"),
            "email": (str, "student email"),
            "course_id": (int, "course ID"),
            "resume": (str, "resume URL"),
            "skills": (str, "comma-separated skills"),
            "enrollment_date": (str, "enrollment date (YYYY-MM-DD)")
        },
        "field_lengths": {
            "name": (3, 100),
            "email": (5, 100),
            "resume": (0, 255),
            "skills": (0, 255)
        }
    }


class SchemaValidator:
    """Validates data structure and format consistency."""
    
    @staticmethod
    def _rows_from_cursor(cursor) -> List[Dict[str, Any]]:
        cols = [d[0] for d in cursor.description]
        rows = cursor.fetchall()
        results: List[Dict[str, Any]] = []
        for r in rows:
            results.append({cols[i]: r[i] for i in range(len(cols))})
        return results

    @staticmethod
    def validate_members_schema(members: List[Dict[str, Any]]) -> SchemaValidationResult:
        """Validate members data structure."""
        errors = []
        warnings = []
        
        schema = SchemaDefinition.MEMBERS_SCHEMA
        
        if not isinstance(members, list):
            return SchemaValidationResult(False, ["Members must be a list"], [])
        
        for idx, member in enumerate(members):
            if not isinstance(member, dict):
                errors.append(f"Member at index {idx} is not a dictionary")
                continue
            
            # Check required fields
            for field in schema["required_fields"]:
                if field not in member:
                    errors.append(f"Member {idx}: Missing required field '{field}'")
                elif member[field] is None:
                    errors.append(f"Member {idx}: Required field '{field}' is None")
            
            # Check field types
            for field, value in member.items():
                if field in schema["field_types"]:
                    expected_type = schema["field_types"][field][0]
                    if not isinstance(value, expected_type):
                        errors.append(
                            f"Member {idx}: Field '{field}' has wrong type. "
                            f"Expected {expected_type.__name__}, got {type(value).__name__}"
                        )
            
            # Check field lengths
            for field, (min_len, max_len) in schema["field_lengths"].items():
                if field in member and member[field] is not None:
                    length = len(str(member[field]))
                    if length < min_len or length > max_len:
                        errors.append(
                            f"Member {idx}: Field '{field}' length {length} "
                            f"not in range [{min_len}, {max_len}]"
                        )
            
            # Check enums
            for field, valid_values in schema["valid_enums"].items():
                if field in member and member[field] not in valid_values:
                    errors.append(
                        f"Member {idx}: Field '{field}' value '{member[field]}' "
                        f"not in valid options: {valid_values}"
                    )
        
        return SchemaValidationResult(len(errors) == 0, errors, warnings)
    
    @staticmethod
    def validate_courses_schema(courses: List[Dict[str, Any]]) -> SchemaValidationResult:
        """Validate courses data structure."""
        errors = []
        warnings = []
        
        schema = SchemaDefinition.COURSES_SCHEMA
        
        if not isinstance(courses, list):
            return SchemaValidationResult(False, ["Courses must be a list"], [])
        
        for idx, course in enumerate(courses):
            if not isinstance(course, dict):
                errors.append(f"Course at index {idx} is not a dictionary")
                continue
            
            # Check required fields
            for field in schema["required_fields"]:
                if field not in course:
                    errors.append(f"Course {idx}: Missing required field '{field}'")
                elif course[field] is None:
                    errors.append(f"Course {idx}: Required field '{field}' is None")
            
            # Check field types
            for field, value in course.items():
                if field in schema["field_types"]:
                    expected_type = schema["field_types"][field][0]
                    if not isinstance(value, expected_type):
                        errors.append(
                            f"Course {idx}: Field '{field}' has wrong type. "
                            f"Expected {expected_type.__name__}, got {type(value).__name__}"
                        )
            
            # Check field lengths
            for field, (min_len, max_len) in schema["field_lengths"].items():
                if field in course and course[field] is not None:
                    length = len(str(course[field]))
                    if length < min_len or length > max_len:
                        errors.append(
                            f"Course {idx}: Field '{field}' length {length} "
                            f"not in range [{min_len}, {max_len}]"
                        )
            
            # Check enums
            for field, valid_values in schema["valid_enums"].items():
                if field in course and course[field] not in valid_values:
                    errors.append(
                        f"Course {idx}: Field '{field}' value '{course[field]}' "
                        f"not in valid options: {valid_values}"
                    )
        
        return SchemaValidationResult(len(errors) == 0, errors, warnings)
    
    @staticmethod
    def validate_students_schema(students: List[Dict[str, Any]]) -> SchemaValidationResult:
        """Validate students data structure."""
        errors = []
        warnings = []
        
        schema = SchemaDefinition.STUDENTS_SCHEMA
        
        if not isinstance(students, list):
            return SchemaValidationResult(False, ["Students must be a list"], [])
        
        for idx, student in enumerate(students):
            if not isinstance(student, dict):
                errors.append(f"Student at index {idx} is not a dictionary")
                continue
            
            # Check required fields
            for field in schema["required_fields"]:
                if field not in student:
                    errors.append(f"Student {idx}: Missing required field '{field}'")
                elif student[field] is None:
                    errors.append(f"Student {idx}: Required field '{field}' is None")
            
            # Check field types
            for field, value in student.items():
                if field in schema["field_types"]:
                    expected_type = schema["field_types"][field][0]
                    if not isinstance(value, expected_type):
                        errors.append(
                            f"Student {idx}: Field '{field}' has wrong type. "
                            f"Expected {expected_type.__name__}, got {type(value).__name__}"
                        )
            
            # Check field lengths
            for field, (min_len, max_len) in schema["field_lengths"].items():
                if field in student and student[field] is not None:
                    length = len(str(student[field]))
                    if length < min_len or length > max_len:
                        errors.append(
                            f"Student {idx}: Field '{field}' length {length} "
                            f"not in range [{min_len}, {max_len}]"
                        )
        
        return SchemaValidationResult(len(errors) == 0, errors, warnings)
    
    @staticmethod
    def validate_json_structure(filepath: str) -> SchemaValidationResult:
        """Validate JSON file structure."""
        errors = []
        warnings = []
        
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
        except FileNotFoundError:
            return SchemaValidationResult(False, [f"File not found: {filepath}"], [])
        except json.JSONDecodeError as e:
            return SchemaValidationResult(False, [f"Invalid JSON: {e}"], [])
        
        # Check top-level structure
        if not isinstance(data, dict):
            errors.append("Root must be a JSON object")
            return SchemaValidationResult(False, errors, warnings)
        
        # Check required keys
        required_keys = ["Members", "Courses", "Students"]
        for key in required_keys:
            if key not in data:
                warnings.append(f"Missing key: '{key}'")
            elif not isinstance(data[key], list):
                errors.append(f"'{key}' must be a list, got {type(data[key]).__name__}")
        
        # Validate each section
        if "Members" in data:
            members_result = SchemaValidator.validate_members_schema(data["Members"])
            errors.extend(members_result.errors)
            warnings.extend(members_result.warnings)
        
        if "Courses" in data:
            courses_result = SchemaValidator.validate_courses_schema(data["Courses"])
            errors.extend(courses_result.errors)
            warnings.extend(courses_result.warnings)
        
        if "Students" in data:
            students_result = SchemaValidator.validate_students_schema(data["Students"])
            errors.extend(students_result.errors)
            warnings.extend(students_result.warnings)
        
        return SchemaValidationResult(len(errors) == 0, errors, warnings)
    
    @staticmethod
    def validate_excel_structure(filepath: str) -> SchemaValidationResult:
        """Validate Excel file structure."""
        try:
            import openpyxl
        except ImportError:
            return SchemaValidationResult(False, ["openpyxl not installed"], [])
        
        errors = []
        warnings = []
        
        try:
            wb = openpyxl.load_workbook(filepath)
        except FileNotFoundError:
            return SchemaValidationResult(False, [f"File not found: {filepath}"], [])
        except Exception as e:
            return SchemaValidationResult(False, [f"Error reading Excel: {e}"], [])
        
        # Check sheet names
        required_sheets = ["Members", "Courses", "Students"]
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                warnings.append(f"Missing sheet: '{sheet}'")
        
        # Validate each sheet
        if "Members" in wb.sheetnames:
            ws = wb["Members"]
            members = SchemaValidator._read_excel_sheet(ws)
            members_result = SchemaValidator.validate_members_schema(members)
            errors.extend(members_result.errors)
            warnings.extend(members_result.warnings)
        
        if "Courses" in wb.sheetnames:
            ws = wb["Courses"]
            courses = SchemaValidator._read_excel_sheet(ws)
            courses_result = SchemaValidator.validate_courses_schema(courses)
            errors.extend(courses_result.errors)
            warnings.extend(courses_result.warnings)
        
        if "Students" in wb.sheetnames:
            ws = wb["Students"]
            students = SchemaValidator._read_excel_sheet(ws)
            students_result = SchemaValidator.validate_students_schema(students)
            errors.extend(students_result.errors)
            warnings.extend(students_result.warnings)
        
        return SchemaValidationResult(len(errors) == 0, errors, warnings)
    
    @staticmethod
    def _read_excel_sheet(ws) -> List[Dict[str, Any]]:
        """Read Excel sheet and convert to list of dicts."""
        rows = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):  # Skip empty rows
                continue
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            rows.append(row_dict)
        
        return rows
    
    @staticmethod
    def print_report(result: SchemaValidationResult) -> None:
        """Print schema validation report."""
        print("\n" + "=" * 70)
        print("SCHEMA VALIDATION REPORT")
        print("=" * 70)
        print(result)

    @staticmethod
    def validate_db_structure(db_path: str) -> SchemaValidationResult:
        """Validate SQLite database structure and row data.

        This checks that required tables exist, required columns are present,
        and then validates rows in the tables against the schema rules.
        """
        errors: List[str] = []
        warnings: List[str] = []

        if not Path(db_path).exists():
            return SchemaValidationResult(False, [f"Database file not found: {db_path}"], [])

        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
        except Exception as e:
            return SchemaValidationResult(False, [f"Unable to open database: {e}"], [])

        try:
            # Discover tables
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [row[0] for row in cursor.fetchall()]

            required_tables = ["Members", "Courses", "Students"]
            for tbl in required_tables:
                if tbl not in tables:
                    warnings.append(f"Missing table: '{tbl}' in database")

            # Validate Members table if present
            if "Members" in tables:
                try:
                    cursor.execute("SELECT * FROM Members;")
                    members = SchemaValidator._rows_from_cursor(cursor)
                    # check columns for required fields
                    cols = [d[0] for d in cursor.description]
                    for field in SchemaDefinition.MEMBERS_SCHEMA["required_fields"]:
                        if field not in cols:
                            errors.append(f"Members table: Missing column '{field}'")
                    members_result = SchemaValidator.validate_members_schema(members)
                    errors.extend(members_result.errors)
                    warnings.extend(members_result.warnings)
                except Exception as e:
                    errors.append(f"Error querying Members table: {e}")

            # Validate Courses table if present
            if "Courses" in tables:
                try:
                    cursor.execute("SELECT * FROM Courses;")
                    courses = SchemaValidator._rows_from_cursor(cursor)
                    cols = [d[0] for d in cursor.description]
                    for field in SchemaDefinition.COURSES_SCHEMA["required_fields"]:
                        if field not in cols:
                            errors.append(f"Courses table: Missing column '{field}'")
                    courses_result = SchemaValidator.validate_courses_schema(courses)
                    errors.extend(courses_result.errors)
                    warnings.extend(courses_result.warnings)
                except Exception as e:
                    errors.append(f"Error querying Courses table: {e}")

            # Validate Students table if present
            if "Students" in tables:
                try:
                    cursor.execute("SELECT * FROM Students;")
                    students = SchemaValidator._rows_from_cursor(cursor)
                    cols = [d[0] for d in cursor.description]
                    for field in SchemaDefinition.STUDENTS_SCHEMA["required_fields"]:
                        if field not in cols:
                            errors.append(f"Students table: Missing column '{field}'")
                    students_result = SchemaValidator.validate_students_schema(students)
                    errors.extend(students_result.errors)
                    warnings.extend(students_result.warnings)
                except Exception as e:
                    errors.append(f"Error querying Students table: {e}")

        finally:
            conn.close()

        return SchemaValidationResult(len(errors) == 0, errors, warnings)


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        
        if filepath.endswith(".json"):
            result = SchemaValidator.validate_json_structure(filepath)
        elif filepath.endswith(".xlsx"):
            result = SchemaValidator.validate_excel_structure(filepath)
        elif filepath.endswith(".db") or filepath.endswith(".sqlite"):
            result = SchemaValidator.validate_db_structure(filepath)
        else:
            print("❌ Unsupported file format. Use .json, .xlsx or .db/.sqlite")
            sys.exit(1)
        
        SchemaValidator.print_report(result)
        sys.exit(0 if result.is_valid else 1)
    else:
        print("Usage: python schema_validator.py <filepath>")
        print("Example: python schema_validator.py tests/test_data/test_data.json")
        sys.exit(1)
